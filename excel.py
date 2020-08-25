import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#Load excel file which you want edit
work_book = xl.load_workbook('transactions.xlsx')

#Select sheet from excel commented one also works
sheet = work_book['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_prize = cell.value * 12
    corrected_prize_cell = sheet.cell(row, 4)
    corrected_prize_cell.value = corrected_prize

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

work_book.save('Transaction3.xlsx')
