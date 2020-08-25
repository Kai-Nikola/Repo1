from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


def result_chart():
    work_book = load_workbook("apprentice.xlsx")

    sheet = work_book.active
    sheet = work_book['Sheet1']

    print("Enter Your Marks")
    hindi = int(input("Hindi : "))
    english = int(input("English: "))
    programming = int(input("Programming: "))
    mathematics = int(input("Mathematics: "))
    science = int(input("Science: "))
    social_science = int(input("social science: "))

    sheet.cell(row=2, column=3).value = hindi
    sheet.cell(row=3, column=3).value = english
    sheet.cell(row=4, column=3).value = programming
    sheet.cell(row=5, column=3).value = mathematics
    sheet.cell(row=6, column=3).value = science
    sheet.cell(row=7, column=3).value = social_science

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=2, max_col=3)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    work_book.save('D:\\NP\\PYTHON\\helloxl\\trial.xlsx')


result_chart()
